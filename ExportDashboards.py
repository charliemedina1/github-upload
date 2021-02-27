#this script exports all customer dashboards to xlsx

import _common.t3_fh_util as util
import os
import datetime
import logging
from openpyxl import Workbook
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
import sys

logging.basicConfig(level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
logger=logging.getLogger()

logger.info("starting...")

#get a list of searches
cfg_CompanyId = sys.argv[1] # pass the company ID in the command line
cfg_FileNameDate = datetime.datetime.utcnow().strftime("%Y%m%d%H%M")

cfg_UseCache = False #this should normally be False unless you are debugging/troubleshooting
cfg_CacheFolder = "cache/" #you may need to manually clear this folder as information may be cached indefinitely




def cache_get(cache_key, cache_key_ext=".json"):
    global cfg_CacheFolder, cfg_UseCache
    file_path=cfg_CacheFolder+ cache_key + cache_key_ext
    if not cfg_UseCache or not os.path.isfile(file_path):
        return False, None
    with open(file_path, 'r') as cache_file:
        return True, util.json.load(cache_file)


def cache_set(cache_key, cache_json, cache_key_ext=".json"):
    global cfg_CacheFolder, cfg_UseCache
    if not cfg_UseCache:
        #caching is disabled
        return
    file_path=cfg_CacheFolder+ cache_key + cache_key_ext
    with open(file_path, 'w') as cache_file:
        util.json.dump(cache_json, cache_file)

#to avoid hammering the servers we cache requests to the local file system
def cacheRequest(cache_key, method,url,headers,gydaToken,json=None):
    global cfg_UseCache
    if cfg_UseCache:
        has_cache,cache_content=cache_get(cache_key)
        if has_cache:
            logger.info("using cached values for %s", cache_key)
            return cache_content
    logger.info("requesting content for %s (cache unavailable; cfg_UseCache=%s)", cache_key, cfg_UseCache)
    cache_content=util.gydaRequest(method=method, url=url, headers=headers, gydaToken=gydaToken, json=json)
    if cfg_UseCache:
        logger.info("caching content for %s", cache_key)
        cache_set(cache_key=cache_key, cache_json=cache_content)
    return cache_content



def add_table(ws, data_rows, display_name, append_no_rows_warning=True, set_col_widths=True, max_col_len=50, min_col_length=5):
    col_qty=len(data_rows[0])
    row_qty=0
    if append_no_rows_warning and len(data_rows)==1:
        data_rows.append(["no data found"])
    max_col_lens=[]
    for row in data_rows:
        row_qty+=1
        col =-1
        if set_col_widths:
            for c in row: #length evaluation 
                col+=1
                l = len(str(c)) + 5
                if l>max_col_len:
                    l=max_col_len
                if l<min_col_length:
                    l=min_col_length
                if col>=len(max_col_lens):
                    max_col_lens.append(l)
                elif max_col_lens[col]<l:
                    max_col_lens[col]=l

        #logger.info("appending row %s - values %s", row_qty, row)
        ws.append(row)
    col_letter=openpyxl.utils.get_column_letter(col_qty)
    tab = Table(displayName=display_name, ref="A1:"+col_letter+str(row_qty))
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    if set_col_widths:
        for i, column_width in enumerate(max_col_lens):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = column_width


def generate_reference_name(name, id_name, id_val, max_len=255):
    id_part = " (%s %s)" % (id_name, id_val)
    reference_name = name
    reference_name = re.compile('[^a-zA-Z 0-9]').sub('_', reference_name)
    current_len=len(reference_name) + len(id_part)
    if current_len>max_len:
        need_to_cut=current_len-max_len
        reference_name=reference_name[0: -1 - need_to_cut]
    reference_name=reference_name + id_part
    if len(reference_name)> max_len:
        raise AssertionError("the sheet name did not correctly stay within the limit of %s %s" % (max_len, sheet_name))
    return reference_name

    

g_GydaToken = util.getFHGydaToken(companyId=cfg_CompanyId)

headers={"referrer":"https://app.meltwater.com/", "origin":"https://app.meltwater.com"}

identity_info = util.gydaRequest(method="get", url="https://v2.walkme.meltwater.io/identity", headers=headers, gydaToken=g_GydaToken)
logger.debug("idenity_info=%s", identity_info)
cfg_CompanyName=identity_info["company"]["name"]

logger.info("running for company id: %s; name: %s", cfg_CompanyId, cfg_CompanyName)

cfg_FileNameRoot = re.compile('[^a-zA-Z0-9]').sub('_', cfg_CompanyName)
cfg_FileNameRoot = re.compile('[_]+').sub('_', cfg_FileNameRoot) #find multiple instances of underscore and replace with single
cfg_FileNameRoot=cfg_FileNameRoot[0:25].lower()
#now push this to a single json object
cfg_FullFileRoot=cfg_FileNameRoot+"_"+cfg_CompanyId+"_"+cfg_FileNameDate


dash_response = cacheRequest(cache_key="%s_dashboard_list" % (cfg_CompanyId), method="get", url="https://app.meltwater.com/dashboard_services/v2/dashboard/list", headers=headers, gydaToken=g_GydaToken)


def dashboardToTable(d, includeHeader=False):
    outr=[]
    if includeHeader:
        outr.append(["Dashboard","Widget #","Attribute","Value 1","Value 2"])
    dcols=["%s (%s)" % (util.jsonField(d,["title"]), util.jsonField(d,["_id"]))]
    
    cts =  util.jsonField(d,["containers"], defaultIfDNE=[])
    widgetCount=0
    for ct in cts:
        containerId=util.jsonField(ct,["id"])
        widgets=util.jsonField(ct,["widgets"], defaultIfDNE=[])
        for w in widgets:
            widgetCount+=1
            inputSeq=0
            wcols=[widgetCount]
            outr.append(dcols + wcols + ["Widget Title & Type", util.jsonField(w,"title"), util.jsonField(w,"type")])
            #add rows for widget start/end dates
            startDate=str(util.jsonField(w,["settings","dateRange","absoluteStartDate"]))
            endDate=str(util.jsonField(w,["settings","dateRange","absoluteEndDate"]))
            outr.append(dcols + wcols + ["Widget Start & End Dates",startDate, endDate])
            ipts = util.jsonField(w,["settings","inputs"],defaultIfDNE=[])
            ipts.extend(util.jsonField(w,["storage","contentStreamInputs"],defaultIfDNE=[]))
            outr.append(dcols + wcols + ["Widget Input Count", len(ipts),"Total inputs in this widget"])
            if len(ipts)==0:
                logger.warn("issuing warning - 0 inputs found for widget %s #%s", dcols, wcols)
                outr.append(dcols + wcols + ["WARNING", "No inputs for this widget"])
            else:
                for i in ipts:
                    inputSeq+=1
                    inputName="%s (%s)" % (util.jsonField(i,"name"),util.jsonField(i,"id"))
                    inputType=util.jsonField(i,"inputType")
                    outr.append(dcols + wcols + ["Widget Input #%s Name & Type" % (inputSeq), inputName, inputType])
    if widgetCount==0:
        logger.warn("issuing warning - no widgets found for dashboard %s", dcols)
        outr.append(dcols + ["n/a","WARNING", "No widgets were found for this dashboard"])
    return outr 
    
    

iDash=0
sheet_rows=[]
for d in dash_response:
    iDash+=1
    logger.info("Dash #%s; dash id = %s; title = %s", iDash, d["_id"], d["title"])
    sheet_rows.extend(dashboardToTable(d, includeHeader=(iDash==1)))
    # if iDash==1:
    #     logger.info("dashjson = %s", d)
    #     sys.exit(0)


if len(sheet_rows)==0:
    raise AssertionError("no dashboards found, ending early")

wb = Workbook()
ws = wb.active
ws.title="Dashboards"

add_table(ws, sheet_rows, "tb_dashboards")


logger.info("saving %s", cfg_FullFileRoot+"_dashboards.xlsx")
wb.save(cfg_FullFileRoot+"_dashboards.xlsx")

logger.info("finished!")
